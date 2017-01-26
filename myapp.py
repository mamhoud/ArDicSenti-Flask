from flask import Flask, flash, redirect, render_template, request, session, abort
import nltk
import os, tempfile, zipfile
import sys
from imp import reload
from io import BytesIO

reload(sys)
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.stem.isri import ISRIStemmer
from nltk.corpus import stopwords
import nltk
import sys
import os.path
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.stem.isri import ISRIStemmer
from nltk.corpus import stopwords

import xlsxwriter
from pandas._testing import isnull
from xlrd import open_workbook
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import random
import pandas as pd


app = Flask(__name__)

@app.route("/app")
def index():
    return render_template(
        '/app/index.html', )

# @app.route("/download_file")
# def download():
#     download_file(request)




# sys.setdefaultencoding('utf-8')


types = {'Root': 0, 'Positive': 1, 'Negative': 2, 'Mis': 3}
data_set = [['Root', 'Positive', 'Negative', 'Mis'], ['حمد', 40, 90, 50], ['وقع', 0, 0, 0], ['لعب', 0, 0, 0],
            ['شرب', 0, 0, 0], ['مشى', 0, 0, 0]]


def stemer(word):
    stem = ISRIStemmer()
    root = stem.stem(word)
    return root


def add_word(iRow, word):
    path = "inter.xlsx"
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name("Sheet1")
    R = ws.cell(row=iRow + 1, column=1)
    P = ws.cell(row=iRow + 1, column=2)
    N = ws.cell(row=iRow + 1, column=3)
    M = ws.cell(row=iRow + 1, column=4)

    R.value = word
    P.value = 0
    N.value = 0
    M.value = 0
    wb.save(path)


# Create a workbook and add a worksheet.
def createFile():
    types = {'Root': 0, 'Positive': 1, 'Negative': 2, 'Mis': 3}
    data_set = [['Root', 'Positive', 'Negative', 'Mis'], ['حمد', 40, 90, 50], ['وقع', 0, 0, 0], ['لعب', 0, 0, 0],
                ['شرب', 0, 0, 0], ['مشى', 0, 0, 0]]
    Row = 0
    Col = 0
    if not os.path.exists('inter.xlsx'):
        workbook = xlsxwriter.Workbook('inter.xlsx')
        worksheet = workbook.add_worksheet()
        # Iterate over the data and write it out row by row.
        for row in data_set:
            for record in row:
                worksheet.write(Row, Col, record)
                Col += 1
            Row += 1
            Col = 0
            workbook.close()
    else:
        return "file exitsts"


def search(word):
    r = 0
    x = 0
    read = open_workbook(r"inter.xlsx")
    for sheet in read.sheets():
        if sheet.name == 'Sheet1':
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    record = sheet.cell(row, col).value
                    if record == word:
                        r = row
                        x = 1
                        break
            if x == 0:
                add_word(sheet.nrows, word)
                r = sheet.nrows
                break

    return r

@app.route("/app/contrib_step1/",methods=['POST'])
def contr_step1():
    read = open_workbook("inter.xlsx")
    data = []

    for sheet in read.sheets():
        if sheet.name == 'Sheet1':
            for row in range(sheet.nrows):
                if row > 0:
                    s = sheet.cell(row, 0).value
                    data.append(s)
    word = random.choice(data)
    return word

@app.route("/app/contribute/",methods=['POST'])
def contr_step_2():
    radiobutton = request.form['addword']
    word = request.form['data']
    row = search(word)

    if radiobutton == "Postive":
        modifyPos(row, 1)
        return render_template( 'app/Positive_added.html')
    elif radiobutton == "Negative":
        modifyNeg(row, 1)
        return render_template( 'app/Negative_added.html')
    elif radiobutton == "Misleading":
        modifyMis(row, 1)
        return render_template('app/Missleading_added.html')


def modifyPos(iRow, newVal):
    path = "inter.xlsx"
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name("Sheet1")
    c = ws.cell(row=iRow + 1, column=types['Positive'] + 1)
    c.value = c.value + newVal
    wb.save(path)


def modifyNeg(iRow, newVal):
    path = "inter.xlsx"
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name("Sheet1")
    c = ws.cell(row=iRow + 1, column=types['Negative'] + 1)
    c.value = c.value + newVal
    wb.save(path)


def modifyMis(iRow, newVal):
    path = "inter.xlsx"
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name("Sheet1")
    c = ws.cell(row=iRow + 1, column=types['Mis'] + 1)
    c.value = c.value + newVal
    wb.save(path)


def Summition(Positive, Negative, Misleading):
    per_array = []
    Sum = Positive + Negative + Misleading
    if Positive > 0:
        Per_pos = (Positive / Sum) * 100
        per_array.append(Per_pos)
    else:
        per_array.append(0)
    if Negative > 0:
        Per_Neg = (Negative / Sum) * 100
        per_array.append(Per_Neg)
    else:
        per_array.append(0)
    if Misleading > 0:
        Per_mis = (Misleading / Sum) * 100
        per_array.append(Per_mis)
    else:
        per_array.append(0)
    return per_array


def file_generator(positive_percentage=0, negative_percentage=0, misleading_percentage=0):
    # If You Want The Final Files As Excel
    # Positive_final = xlsxwriter.Workbook('Positive_final.xlsx')
    # worksheet_p = Positive_final.add_worksheet()
    # Negative_final = xlsxwriter.Workbook('Negative_final.xlsx')
    # worksheet_n = Negative_final.add_worksheet()
    # Misleading_final = xlsxwriter.Workbook('Misleading_final.xlsx')
    # worksheet_m = Misleading_final.add_worksheet()


    read = open_workbook("inter.xlsx")
    # read_pos = open_workbook("Positive_final.xlsx")
    # read_neg = open_workbook("Negative_final.xlsx")
    # read_mis = open_workbook("Misleading_final.xlsx")

    data_pos = []
    data_neg = []
    data_mis = []

    for sheet in read.sheets():
        if sheet.name == 'Sheet1':
            for row in range(sheet.nrows):
                if row > 0:
                    s = sheet.cell(row, 0).value
                    s1 = sheet.cell(row, 1).value
                    s2 = sheet.cell(row, 2).value
                    s3 = sheet.cell(row, 3).value
                    row_array = Summition(s1, s2, s3)

                    # worksheet_p.write(0, 0, "Root")
                    # worksheet_n.write(0, 0, "Root")
                    # worksheet_m.write(0, 0, "Root")

                    if (int(row_array[0]) >= positive_percentage) and (int(row_array[0]) != 0) and (
                        positive_percentage != 0):
                        data_pos.append(s)

                    if (int(row_array[1]) >= negative_percentage) and (int(row_array[1]) != 0) and (
                        negative_percentage != 0):
                        data_neg.append(s)

                    if (int(row_array[2]) >= misleading_percentage) and (int(row_array[2]) != 0) and (
                        misleading_percentage != 0):
                        data_mis.append(s)

    pf = pd.DataFrame(data_pos, columns=["Root"])
    pf.to_csv('Final_positive.csv', index=False)

    nf = pd.DataFrame(data_neg, columns=["Root"])
    nf.to_csv('Final_negative.csv', index=False)

    mf = pd.DataFrame(data_mis, columns=["Root"])
    mf.to_csv('Final_misleading.csv', index=False)

    # If You Want The Final Files As Excel
    # for i, e in enumerate(data_pos):
    #     worksheet_p.write(i, 0, e)
    # for i, e in enumerate(data_neg):
    #     worksheet_n.write(i, 0, e)
    # for i, e in enumerate(data_mis):
    #     worksheet_m.write(i, 0, e)
    #
    # Positive_final.close()
    # Negative_final.close()
    # Misleading_final.close()

@app.route('/app/addword/', methods=['POST'])
def addword():
    word = request.form['word']
    radiobutton = request.form['addword']
    createFile()

    if word == "" and radiobutton == "":
        return render_template( 'app/Nothing_added.html')
    else:
        root = stemer(word)
        row = search(root)
        if radiobutton == "po":
            modifyPos(row, 1)
            return render_template('app/Positive_added.html')
        elif radiobutton == "ne":
            modifyNeg(row, 1)
            return render_template('app/Negative_added.html')
        elif radiobutton == "mis":
            modifyMis(row, 1)
            return render_template('app/Missleading_added.html')


# def zip_file(request, filenames):
#     filenames = filenames
#     zip_subdir = "somefiles"
#     zip_filename = "%s.zip" % zip_subdir
#     s = BytesIO()
#     zf = zipfile.ZipFile(s, "w")
#     for fpath in filenames:
#         fdir, fname = os.path.split(fpath)
#         zip_path = os.path.join(zip_subdir, fname)
#         zf.write(fpath, zip_path)
#     zf.close()
#     resp = HttpResponse(s.getvalue(), content_type="application/x-zip-compressed")
#     resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename
#     return resp


# def download_file(request):
#     pos_percentage = request.POST['text-positive']
#     neg_percentage = request.POST['text-negative']
#     mis_percentage = request.POST['text-misleading']
#     value = request.POST.getlist("optradio")
#     # if (pos_percentage != "" or neg_percentage != "" or mis_percentage != ""):
#     #     file_generator(pos_percentage,neg_percentage,mis_percentage)
#
#     if len(value) == 3:
#
#         if value[0] == 'postive' and value[1] == 'negative' and value[
#             2] == "miss" and pos_percentage != "" and neg_percentage != "" and mis_percentage != "":
#             file_generator(int(pos_percentage), int(neg_percentage), int(mis_percentage))
#             filenames = ["Final_positive.csv", "Final_negative.csv", "Final_misleading.csv"]
#             return zip_file(request, filenames)
#
#
#
#     elif len(value) == 2:
#
#         if value[0] == 'postive' and value[1] == 'negative' and pos_percentage != "" and neg_percentage != "":
#             file_generator(int(pos_percentage), int(neg_percentage))
#             filenames = ["Final_positive.csv", "Final_negative.csv"]
#             return zip_file(request, filenames)
#
#         elif value[0] == 'postive' and value[1] == 'miss' and pos_percentage != "" and mis_percentage != "":
#             file_generator(int(pos_percentage), 0, int(mis_percentage))
#             filenames = ["Final_positive.csv", "Final_misleading.csv"]
#
#             return zip_file(request, filenames)
#
#         elif value[0] == 'negative' and value[1] == 'miss' and neg_percentage != "" and mis_percentage != "":
#             file_generator(0, int(neg_percentage), int(mis_percentage))
#             filenames = ["Final_negative.csv", "Final_misleading.csv"]
#             return zip_file(request, filenames)
#
#
#
#     elif len(value) == 1:
#
#         if value[0] == 'postive' and pos_percentage != "":
#             file_generator(int(pos_percentage))
#             filenames = ["Final_positive.csv"]
#             return zip_file(request, filenames)
#
#
#         elif value[0] == 'negative' and neg_percentage != "":
#             file_generator(0, int(neg_percentage))
#             filenames = ["Final_negative.csv"]
#             return zip_file(request, filenames)
#
#         elif value[0] == 'miss' and mis_percentage != "":
#             file_generator(0, 0, int(mis_percentage))
#             filenames = ["Final_misleading.csv"]
#
#             return zip_file(request, filenames)
#
#
if __name__ == "__main__":
    app.run()









